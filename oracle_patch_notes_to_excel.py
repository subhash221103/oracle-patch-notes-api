#!/usr/bin/env python3
"""
oracle_patch_notes_to_excel.py
================================

Fetches Oracle Fusion Cloud readiness / "What's New" content from Oracle Help
Center and writes it into an Excel (.xlsx) workbook — one sheet per module.

USAGE
-----
    # All modules, all pillars for a release:
    python oracle_patch_notes_to_excel.py --release 26b --all

    # Single module:
    python oracle_patch_notes_to_excel.py --release 26b --url scm/26b/mfg26b/index.html

    # List discovered modules for a release:
    python oracle_patch_notes_to_excel.py --release 26b --list-modules

INSTALL
-------
    pip install httpx beautifulsoup4 openpyxl
"""

import argparse
import logging
import re
import sys
from typing import Optional

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger("oracle-patch-notes-xlsx")

BASE = "https://docs.oracle.com/en/cloud/saas/readiness"

# Pillars to scan when --all is used
PILLAR_ALL_PAGES = ["scm", "hcm", "erp"]

MAX_PAGES_DEFAULT = 60

# Friendly name -> URL path (relative to BASE)
MODULE_NAME_MAP: dict[str, str] = {
    # SCM
    "manufacturing":                    "scm/26b/mfg26b/index.html",
    "inventory":                        "scm/26b/inv26b/index.html",
    "procurement":                      "scm/26b/proc26b/index.html",
    "self-service-procurement":         "scm/26b/ssproc26b/index.html",
    "maintenance":                      "scm/26b/maint26b/index.html",
    "plm":                              "scm/26b/plm26b/index.html",
    "order-management":                 "scm/26b/order26b/index.html",
    "supply-planning":                  "scm/26b/scp26b/index.html",
    "demand-management":                "scm/26b/demand26b/index.html",
    "sales-and-operations-planning":    "scm/26b/sop26b/index.html",
    "supply-chain-collaboration":       "scm/26b/sccv26b/index.html",
    # Logistics
    "warehouse-management":             "logistics/26b/wms26b/index.html",
    "transportation-management":        "logistics/26b/otm26b/index.html",
    "global-trade-management":          "logistics/26b/gtm26b/index.html",
    # Common
    "common":                           "common/26b/common26b/index.html",
    # HCM
    "benefits":                         "hcm/26b/benf-26b/index.html",
    "compensation":                     "hcm/26b/comp-26b/index.html",
    "hr":                               "hcm/26b/hcom-26b/index.html",
    "hr-core":                          "hcm/26b/hcom-26b/index.html",
    "union-management":                 "hcm/26b/hure-26b/index.html",
    "workforce-management":             "hcm/26b/opma-26b/index.html",
    "dynamic-skills":                   "hcm/26b/dyns-26b/index.html",
    "learning":                         "hcm/26b/lear-26b/index.html",
    "me-experience":                    "hcm/26b/meex-26b/index.html",
    "recruiting":                       "hcm/26b/recr-26b/index.html",
    "talent-management":                "hcm/26b/tama-26b/index.html",
    "absence-management":               "hcm/26b/amg-26b/index.html",
    "time-labor":                       "hcm/26b/tila-26b/index.html",
    "health-safety":                    "hcm/26b/wohs-26b/index.html",
    "workforce-modeling":               "hcm/26b/wosc-26b/index.html",
    "payroll":                          "hcm/26b/payr-26b/index.html",
    # ERP / Financials
    "financials":                       "erp/26b/fins26b/index.html",
    "self-service-financials":          "erp/26b/ssfin26b/index.html",
    "project-management":               "erp/26b/ppm26b/index.html",
    "risk-management":                  "erp/26b/risk26b/index.html",
}

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

def discover_modules(release: str) -> list[tuple[str, str]]:
    """
    Scrape pillar-all pages and return a deduplicated list of
    (index_url, sheet_label) for the given release.
    """
    release = release.strip().lower()
    seen_urls = set()
    results = []

    for pillar in PILLAR_ALL_PAGES:
        url = f"{BASE}/{pillar}-all.html"
        try:
            r = httpx.get(url, timeout=15, follow_redirects=True,
                          headers={"User-Agent": "Mozilla/5.0"})
            r.raise_for_status()
        except httpx.HTTPError as e:
            logger.warning("Could not fetch %s: %s", url, e)
            continue

        soup = BeautifulSoup(r.text, "html.parser")
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if release in href.lower() and href.endswith("index.html"):
                full_url = href if href.startswith("http") else f"{BASE}/{href}"
                if full_url not in seen_urls:
                    seen_urls.add(full_url)
                    # Derive a short sheet label from the slug portion
                    parts = href.rstrip("/").replace("index.html", "").strip("/").split("/")
                    slug = parts[-1] if parts else href
                    results.append((full_url, slug))

    return results


def get_module_title(index_url: str) -> str:
    """Fetch the index page and return its <title> text as a sheet name."""
    try:
        html = _fetch_page(index_url)
        soup = BeautifulSoup(html, "html.parser")
        title_tag = soup.find("title")
        if title_tag:
            t = title_tag.get_text(strip=True)
            # Strip release suffix like "26B What's New"
            t = re.sub(r"\s*\d+[A-Z]\s+What'?s?\s+New\s*$", "", t, flags=re.I).strip()
            t = re.sub(r"\s*What'?s?\s+New\s*$", "", t, flags=re.I).strip()
            return t
    except Exception:
        pass
    return index_url.split("/")[-2]  # fallback: slug


# ---------------------------------------------------------------------------
# Fetching + parsing
# ---------------------------------------------------------------------------

def _fetch_page(url: str) -> str:
    logger.info("Fetching %s", url)
    resp = httpx.get(url, timeout=30, follow_redirects=True,
                     headers={"User-Agent": "Mozilla/5.0 (compatible; OraclePatchNotesXLSX/1.0)"})
    resp.raise_for_status()
    return resp.text


def _next_link(html: str, base_url: str) -> Optional[str]:
    m = re.search(r'href="([^"]+)"[^>]*>\s*Next\s*<', html, re.I)
    if not m:
        return None
    nxt = m.group(1)
    return nxt if nxt.startswith("http") else str(httpx.URL(base_url).join(nxt))


def fetch_all_pages(index_url: str, max_pages: int = MAX_PAGES_DEFAULT) -> list[tuple[str, str]]:
    """Return (url, html) tuples for the index page plus each 'Next' page."""
    pages = []
    url = index_url
    seen = set()
    for _ in range(max_pages):
        if not url or url in seen:
            break
        seen.add(url)
        try:
            html = _fetch_page(url)
        except httpx.HTTPError as e:
            logger.error("Fetch error for %s: %s", url, e)
            break
        pages.append((url, html))
        url = _next_link(html, url)
    return pages


def _parse_feature_summary(pages: list[tuple[str, str]]) -> dict[str, str]:
    """Return mapping of feature page filename -> work area from the Feature Summary table."""
    work_area_map = {}
    for url, html in pages:
        soup = BeautifulSoup(html, "html.parser")
        h1 = soup.find("h1")
        if not h1 or "feature summary" not in h1.get_text(strip=True).lower():
            continue
        table = soup.find("table", class_="rfs_table")
        if not table:
            continue
        current_module = ""
        for row in table.find_all("tr"):
            cells = row.find_all("td")
            if len(cells) < 2:
                continue
            module_text = cells[0].get_text(strip=True)
            if module_text:
                current_module = module_text
            link = cells[1].find("a", href=True)
            if link:
                href = link["href"].split("#")[0]
                work_area_map[href] = current_module
    return work_area_map


SKIP_TITLES = {
    "revision history", "feature summary", "overview",
    "pre-update and post-update tasks",
    "optional uptake of new features (opt in)",
    "important actions and considerations",
    "opt in expiration",
}

SECTION_HEADINGS = {
    "steps to enable and configure": "steps_to_enable",
    "tips and considerations": "tips",
    "key resources": "key_resources",
    "access requirements": "access_requirements",
}

AI_KEYWORDS = re.compile(r'\bai\s*agent\b|\bagentic\b', re.I)
AGENT_WORK_AREA_KW = re.compile(r'\bagent\b', re.I)
REDWOOD_KW = re.compile(r'\bredwood\b', re.I)


def parse_features(pages: list[tuple[str, str]]) -> list[dict]:
    work_area_map = _parse_feature_summary(pages)
    rows = []

    for url, html in pages:
        soup = BeautifulSoup(html, "html.parser")
        for tag in soup(["script", "style"]):
            tag.decompose()
        for tag in soup.find_all("div", class_="noscript"):
            tag.decompose()
        for tag in soup.find_all("div", id="copyright"):
            tag.decompose()

        h1 = soup.find("h1")
        if not h1:
            continue
        feature_title = h1.get_text(strip=True)
        if feature_title.strip().lower() in SKIP_TITLES:
            continue

        page_file = url.split("/")[-1].split("#")[0]
        work_area = work_area_map.get(page_file, "")

        row = {
            "feature": feature_title,
            "work_area": work_area,
            "description": "",
            "steps_to_enable": "",
            "tips": "",
            "key_resources": "",
            "access_requirements": "",
            "is_ai_agent": bool(AI_KEYWORDS.search(feature_title) or AGENT_WORK_AREA_KW.search(work_area)),
            "is_redwood": bool(REDWOOD_KW.search(feature_title)),
        }

        current_bucket = "description"
        buffer: list[str] = []

        def flush(bucket):
            if buffer:
                row[bucket] = (row[bucket] + " " + " ".join(buffer)).strip()
                buffer.clear()

        for el in h1.find_all_next(["h1", "h2", "p", "li"]):
            if el.name == "h1":
                break
            text = el.get_text(strip=True)
            if not text:
                continue
            if el.name == "h2":
                flush(current_bucket)
                current_bucket = SECTION_HEADINGS.get(text.strip().lower(), "description")
            else:
                buffer.append(text)
        flush(current_bucket)

        rows.append(row)

    return rows


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

HEADERS = ["Sr No.", "Release", "Work Area (Module)", "Feature Description",
           "Oracle AI Agent Feature (Y/N)", "Redwood Feature (Y/N)", "Feature Details"]
COL_WIDTHS = [8, 10, 32, 55, 24, 20, 80]


def write_module_sheet(wb: Workbook, sheet_name: str, release: str, rows: list[dict]):
    sheet = wb.create_sheet(sheet_name[:31])

    for col, (title, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = sheet.cell(row=1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        from openpyxl.utils import get_column_letter
        sheet.column_dimensions[get_column_letter(col)].width = width

    sheet.freeze_panes = "A2"

    for i, row in enumerate(rows, start=1):
        feature_details_parts = [row["description"]]
        if row["steps_to_enable"]:
            feature_details_parts.append("Steps to Enable: " + row["steps_to_enable"])
        if row["tips"]:
            feature_details_parts.append("Tips: " + row["tips"])
        feature_details = "\n\n".join(p for p in feature_details_parts if p)

        values = [
            i,
            release.upper(),
            row["work_area"],
            row["feature"],
            "Y" if row["is_ai_agent"] else "N",
            "Y" if row["is_redwood"] else "N",
            feature_details,
        ]
        for c, val in enumerate(values, start=1):
            cell = sheet.cell(row=i + 1, column=c, value=val)
            cell.font = BODY_FONT
            cell.alignment = WRAP

    if not rows:
        sheet.cell(row=2, column=1, value="No features found.")


def build_summary_sheet(wb: Workbook, release: str, summary_rows: list[tuple]):
    summary = wb.create_sheet("Summary", 0)
    summary["A1"] = f"Oracle Fusion Cloud Readiness — Release: {release.upper()}"
    summary["A1"].font = Font(name="Arial", bold=True, size=13)
    summary.merge_cells("A1:D1")

    for col, title in enumerate(["Module", "Features", "Sheet", "Source URL"], start=1):
        cell = summary.cell(row=3, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    for i, row in enumerate(summary_rows, start=4):
        for c, val in enumerate(row, start=1):
            summary.cell(row=i, column=c, value=val).font = BODY_FONT

    summary.column_dimensions["A"].width = 45
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 32
    summary.column_dimensions["D"].width = 75


# ---------------------------------------------------------------------------
# Main logic
# ---------------------------------------------------------------------------

def process_module(wb: Workbook, index_url: str, sheet_label: str,
                   release: str, max_pages: int) -> tuple:
    """Fetch, parse, write one module. Returns a summary_rows entry."""
    title = get_module_title(index_url)
    sheet_name = (title or sheet_label)[:31]

    try:
        pages = fetch_all_pages(index_url, max_pages=max_pages)
        rows = parse_features(pages)
    except httpx.HTTPError as e:
        logger.error("Failed: %s — %s", index_url, e)
        return (title, f"FAILED: {e}", sheet_name, index_url)

    write_module_sheet(wb, sheet_name, release, rows)
    logger.info("Module '%s': %d features", sheet_name, len(rows))
    return (title, len(rows), sheet_name, index_url)


def main():
    parser = argparse.ArgumentParser(
        description="Fetch Oracle Fusion Cloud readiness content and export to Excel."
    )
    parser.add_argument("--release", required=True, help="Release code, e.g. 26b, 26a")
    parser.add_argument("--all", action="store_true",
                        help="Fetch all modules across all pillars for the release")
    parser.add_argument("--module", help="Module name, e.g. manufacturing, financials, payroll")
    parser.add_argument("--url", help="Single module index URL path, e.g. scm/26b/mfg26b/index.html")
    parser.add_argument("--list-modules", action="store_true",
                        help="List known module names and exit")
    parser.add_argument("--max-pages", type=int, default=MAX_PAGES_DEFAULT)
    parser.add_argument("--out", help="Output .xlsx filename")

    args = parser.parse_args()

    if args.list_modules:
        print(f"Known module names (for --module). Replace '26b' with your release:\n")
        for name, path in sorted(MODULE_NAME_MAP.items()):
            print(f"  {name:35s} -> {path}")
        return

    wb = Workbook()
    wb.remove(wb.active)
    summary_rows = []
    release = args.release.strip().lower()

    if args.all:
        modules = discover_modules(release)
        if not modules:
            print("No modules discovered. Check the release code or network access.")
            sys.exit(1)
        print(f"Discovered {len(modules)} modules. Fetching...")
        for index_url, slug in modules:
            entry = process_module(wb, index_url, slug, release, args.max_pages)
            summary_rows.append(entry)
    elif args.module:
        name = args.module.strip().lower()
        path = MODULE_NAME_MAP.get(name)
        if not path:
            close = [k for k in MODULE_NAME_MAP if name in k or k in name]
            hint = f" Did you mean: {', '.join(close)}?" if close else " Use --list-modules to see all names."
            parser.error(f"Unknown module '{args.module}'.{hint}")
        # Substitute the release in the path (paths are stored with 26b as default)
        path = re.sub(r'26b', release, path, flags=re.I)
        index_url = f"{BASE}/{path}"
        entry = process_module(wb, index_url, name, release, args.max_pages)
        summary_rows.append(entry)
    elif args.url:
        index_url = args.url if args.url.startswith("http") else f"{BASE}/{args.url}"
        entry = process_module(wb, index_url, args.url.split("/")[-2], release, args.max_pages)
        summary_rows.append(entry)
    else:
        parser.error("Provide --all, --module <name>, or --url <path>")

    build_summary_sheet(wb, release, summary_rows)

    out_path = args.out or f"oracle_fusion_{release}_readiness.xlsx"
    wb.save(out_path)
    total = sum(r[1] for r in summary_rows if isinstance(r[1], int))
    print(f"\nSaved: {out_path}  ({len(summary_rows)} modules, {total} total features)")


if __name__ == "__main__":
    sys.exit(main() or 0)
